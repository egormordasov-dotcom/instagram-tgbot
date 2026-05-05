from aiogram import Router, F
from aiogram.types import CallbackQuery, InlineKeyboardMarkup, InlineKeyboardButton
from datetime import datetime, timezone, timedelta
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

router = Router()

def reports_menu():
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="📊 Отчёт за период",    callback_data="rep_period")],
        [InlineKeyboardButton(text="📈 Прирост просмотров", callback_data="rep_growth")],
        [InlineKeyboardButton(text="◀️ Назад",             callback_data="menu_main")],
    ])

def period_picker(report_type: str):
    """Кнопки выбора периода."""
    now = datetime.now()
    buttons = []
    for i in range(6):
        month = (now.month - i - 1) % 12 + 1
        year  = now.year if now.month - i > 0 else now.year - 1
        label = datetime(year, month, 1).strftime("%B %Y")
        data  = f"{report_type}|{year}|{month}"
        buttons.append([InlineKeyboardButton(text=label, callback_data=data)])
    buttons.append([InlineKeyboardButton(text="◀️ Назад", callback_data="menu_reports")])
    return InlineKeyboardMarkup(inline_keyboard=buttons)

@router.callback_query(F.data == "menu_reports")
async def menu_reports(call: CallbackQuery):
    await call.message.edit_text("📊 Отчёты по артикулам:", reply_markup=reports_menu())

@router.callback_query(F.data == "rep_period")
async def rep_period(call: CallbackQuery):
    await call.message.edit_text(
        "📊 <b>Отчёт за период</b>\n\nВыберите месяц:",
        parse_mode="HTML",
        reply_markup=period_picker("rp")
    )

@router.callback_query(F.data == "rep_growth")
async def rep_growth(call: CallbackQuery):
    await call.message.edit_text(
        "📈 <b>Прирост просмотров</b>\n\nВыберите месяц:",
        parse_mode="HTML",
        reply_markup=period_picker("rg")
    )

@router.callback_query(F.data.startswith("rp|"))
async def handle_period_report(call: CallbackQuery, pool):
    _, year, month = call.data.split("|")
    year, month = int(year), int(month)
    await call.message.edit_text("⏳ Формируем отчёт...")

    user_id   = call.from_user.id
    date_from = datetime(year, month, 1, tzinfo=timezone.utc)
    next_month = month % 12 + 1
    next_year  = year + (1 if month == 12 else 0)
    date_to   = datetime(next_year, next_month, 1, tzinfo=timezone.utc)

    text, excel = await build_period_report(pool, user_id, date_from, date_to)

    await call.message.edit_text(text, parse_mode="HTML", reply_markup=InlineKeyboardMarkup(
        inline_keyboard=[[InlineKeyboardButton(text="◀️ К отчётам", callback_data="menu_reports")]]
    ))
    if excel:
        month_str = date_from.strftime("%Y_%m")
        await call.message.answer_document(
            document=excel,
            caption=f"📊 Отчёт за {date_from.strftime('%B %Y')}"
        )

@router.callback_query(F.data.startswith("rg|"))
async def handle_growth_report(call: CallbackQuery, pool):
    _, year, month = call.data.split("|")
    year, month = int(year), int(month)
    await call.message.edit_text("⏳ Считаем прирост...")

    user_id   = call.from_user.id
    date_from = datetime(year, month, 1, tzinfo=timezone.utc)
    next_month = month % 12 + 1
    next_year  = year + (1 if month == 12 else 0)
    date_to   = datetime(next_year, next_month, 1, tzinfo=timezone.utc)

    text, excel = await build_growth_report(pool, user_id, date_from, date_to)

    await call.message.edit_text(text, parse_mode="HTML", reply_markup=InlineKeyboardMarkup(
        inline_keyboard=[[InlineKeyboardButton(text="◀️ К отчётам", callback_data="menu_reports")]]
    ))
    if excel:
        await call.message.answer_document(
            document=excel,
            caption=f"📈 Прирост за {date_from.strftime('%B %Y')}"
        )

async def build_period_report(pool, user_id: int, date_from, date_to):
    """Отчёт за период: видео вышедшие в этот период + их текущие просмотры."""
    async with pool.acquire() as conn:
        rows = await conn.fetch("""
            SELECT
                a.article,
                COUNT(DISTINCT v.id) AS video_count,
                COALESCE(SUM(v.views), 0) AS total_views,
                STRING_AGG(DISTINCT acc.username, ', ') AS accounts
            FROM articles a
            JOIN article_matches am ON am.article_id = a.id
            JOIN videos v  ON v.id = am.video_id
            JOIN accounts acc ON acc.id = v.account_id
            WHERE a.user_id = $1
              AND v.published_at >= $2
              AND v.published_at <  $3
            GROUP BY a.article
            ORDER BY total_views DESC
        """, user_id, date_from, date_to)

    if not rows:
        return (
            f"📊 За {date_from.strftime('%B %Y')} видео с вашими артикулами не найдено.\n\n"
            "<i>Убедитесь что аккаунты синхронизированы и артикулы добавлены.</i>",
            None
        )

    total_videos = sum(r["video_count"] for r in rows)
    total_views  = sum(r["total_views"] for r in rows)

    lines = [
        f"📊 <b>Отчёт за {date_from.strftime('%B %Y')}</b>\n",
        f"Видео: {total_videos} | Просмотры: {total_views:,}\n",
    ]
    for r in rows[:15]:
        lines.append(
            f"• <code>{r['article']}</code> — "
            f"{r['video_count']} видео, {r['total_views']:,} просмотров"
        )
    if len(rows) > 15:
        lines.append(f"\n<i>...и ещё {len(rows)-15} артикулов — смотрите в Excel</i>")

    excel = build_excel_period(rows, date_from)
    return "\n".join(lines), excel

async def build_growth_report(pool, user_id: int, date_from, date_to):
    """Прирост просмотров за период по всем видео с артикулом."""
    async with pool.acquire() as conn:
        rows = await conn.fetch("""
            WITH
            snap_end AS (
                SELECT DISTINCT ON (s.video_id)
                    s.video_id, s.views
                FROM snapshots s
                WHERE s.taken_at <= $3
                ORDER BY s.video_id, s.taken_at DESC
            ),
            snap_start AS (
                SELECT DISTINCT ON (s.video_id)
                    s.video_id, s.views
                FROM snapshots s
                WHERE s.taken_at < $2
                ORDER BY s.video_id, s.taken_at DESC
            )
            SELECT
                a.article,
                COUNT(DISTINCT v.id) AS video_count,
                COALESCE(SUM(se.views - COALESCE(ss.views, 0)), 0) AS growth
            FROM articles a
            JOIN article_matches am ON am.article_id = a.id
            JOIN videos v ON v.id = am.video_id
            LEFT JOIN snap_end   se ON se.video_id = v.id
            LEFT JOIN snap_start ss ON ss.video_id = v.id
            WHERE a.user_id = $1
            GROUP BY a.article
            HAVING COALESCE(SUM(se.views - COALESCE(ss.views, 0)), 0) > 0
            ORDER BY growth DESC
        """, user_id, date_from, date_to)

    if not rows:
        return (
            f"📈 За {date_from.strftime('%B %Y')} прирост не зафиксирован.\n\n"
            "<i>Нужно хотя бы два замера просмотров в этом периоде.</i>",
            None
        )

    total_growth = sum(r["growth"] for r in rows)
    lines = [
        f"📈 <b>Прирост за {date_from.strftime('%B %Y')}</b>\n",
        f"Суммарный прирост: {total_growth:,} просмотров\n",
    ]
    for r in rows[:15]:
        lines.append(
            f"• <code>{r['article']}</code> — "
            f"+{r['growth']:,} просмотров ({r['video_count']} видео)"
        )
    if len(rows) > 15:
        lines.append(f"\n<i>...и ещё {len(rows)-15} — смотрите в Excel</i>")

    excel = build_excel_growth(rows, date_from)
    return "\n".join(lines), excel

async def build_summary_text(pool, user_id: int, period: str) -> str:
    """Для авто-отчётов."""
    now = datetime.now(tz=timezone.utc)
    if period == "day":
        date_from = now - timedelta(days=1)
        label = "за последние сутки"
    elif period == "week":
        date_from = now - timedelta(weeks=1)
        label = "за последнюю неделю"
    else:
        date_from = now.replace(day=1, hour=0, minute=0, second=0)
        label = f"за {now.strftime('%B')}"

    async with pool.acquire() as conn:
        rows = await conn.fetch("""
            SELECT acc.username,
                COUNT(DISTINCT v.id) as videos,
                COALESCE(SUM(v.views), 0) as views
            FROM accounts acc
            JOIN videos v ON v.account_id = acc.id
            WHERE acc.user_id=$1 AND v.published_at >= $2
            GROUP BY acc.username ORDER BY views DESC
        """, user_id, date_from)

    if not rows:
        return f"🔔 Авто-отчёт {label}:\nВидео не найдено."

    lines = [f"🔔 <b>Авто-отчёт {label}</b>\n"]
    for r in rows:
        lines.append(f"@{r['username']}: {r['videos']} видео, {r['views']:,} просмотров")
    return "\n".join(lines)

def build_excel_period(rows, date_from) -> io.BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Отчёт"

    hfill = PatternFill("solid", fgColor="2D6A9F")
    hfont = Font(bold=True, color="FFFFFF")
    ctr   = Alignment(horizontal="center")

    headers = ["Артикул", "Видео", "Просмотры", "Аккаунты"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font=hfont; cell.fill=hfill; cell.alignment=ctr

    for ri, r in enumerate(rows, 2):
        ws.cell(row=ri, column=1, value=r["article"])
        ws.cell(row=ri, column=2, value=r["video_count"])
        ws.cell(row=ri, column=3, value=r["total_views"])
        ws.cell(row=ri, column=4, value=r["accounts"])

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["C"].width = 15

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    buf.name = f"report_{date_from.strftime('%Y_%m')}.xlsx"
    return buf

def build_excel_growth(rows, date_from) -> io.BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Прирост"

    hfill = PatternFill("solid", fgColor="1E6E42")
    hfont = Font(bold=True, color="FFFFFF")
    ctr   = Alignment(horizontal="center")

    headers = ["Артикул", "Видео", "Прирост просмотров"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font=hfont; cell.fill=hfill; cell.alignment=ctr

    for ri, r in enumerate(rows, 2):
        ws.cell(row=ri, column=1, value=r["article"])
        ws.cell(row=ri, column=2, value=r["video_count"])
        ws.cell(row=ri, column=3, value=r["growth"])

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["C"].width = 20

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    buf.name = f"growth_{date_from.strftime('%Y_%m')}.xlsx"
    return buf
