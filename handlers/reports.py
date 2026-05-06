from aiogram import Router, F
from aiogram.types import CallbackQuery, Message, InlineKeyboardMarkup, InlineKeyboardButton
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from datetime import datetime, timezone, timedelta
import calendar
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

router = Router()

class ReportStates(StatesGroup):
    picking_start = State()
    picking_end   = State()

# ── Меню отчётов ─────────────────────────────────────────────
def reports_menu():
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="📊 Отчёт за период",    callback_data="rep_period")],
        [InlineKeyboardButton(text="📈 Прирост просмотров", callback_data="rep_growth")],
        [InlineKeyboardButton(text="◀️ Назад",             callback_data="menu_main")],
    ])

@router.callback_query(F.data == "menu_reports")
async def menu_reports(call: CallbackQuery):
    await call.message.edit_text("📊 Отчёты по артикулам:", reply_markup=reports_menu())

# ── Календарь ────────────────────────────────────────────────
def make_calendar(year: int, month: int, prefix: str, selected: str = None):
    """Строит инлайн-календарь для выбора даты."""
    now = datetime.now()
    month_name = datetime(year, month, 1).strftime("%B %Y")
    _, days_in_month = calendar.monthrange(year, month)

    rows = []

    # Заголовок с навигацией по месяцам
    prev_month = month - 1 if month > 1 else 12
    prev_year  = year if month > 1 else year - 1
    next_month = month + 1 if month < 12 else 1
    next_year  = year if month < 12 else year + 1

    rows.append([
        InlineKeyboardButton(text="◀", callback_data=f"{prefix}_nav_{prev_year}_{prev_month}"),
        InlineKeyboardButton(text=month_name, callback_data="ignore"),
        InlineKeyboardButton(text="▶", callback_data=f"{prefix}_nav_{next_year}_{next_month}"),
    ])

    # Дни недели
    rows.append([
        InlineKeyboardButton(text=d, callback_data="ignore")
        for d in ["Пн","Вт","Ср","Чт","Пт","Сб","Вс"]
    ])

    # Дни месяца
    first_weekday = datetime(year, month, 1).weekday()
    week = [InlineKeyboardButton(text=" ", callback_data="ignore")] * first_weekday

    for day in range(1, days_in_month + 1):
        date_str = f"{year}-{month:02d}-{day:02d}"
        is_future = datetime(year, month, day) > now
        if is_future:
            week.append(InlineKeyboardButton(text="·", callback_data="ignore"))
        elif date_str == selected:
            week.append(InlineKeyboardButton(text=f"✅{day}", callback_data=f"{prefix}_day_{date_str}"))
        else:
            week.append(InlineKeyboardButton(text=str(day), callback_data=f"{prefix}_day_{date_str}"))

        if len(week) == 7:
            rows.append(week)
            week = []

    if week:
        while len(week) < 7:
            week.append(InlineKeyboardButton(text=" ", callback_data="ignore"))
        rows.append(week)

    # Быстрые периоды
    rows.append([
        InlineKeyboardButton(text="7 дней",  callback_data=f"{prefix}_quick_7"),
        InlineKeyboardButton(text="30 дней", callback_data=f"{prefix}_quick_30"),
        InlineKeyboardButton(text="90 дней", callback_data=f"{prefix}_quick_90"),
    ])
    rows.append([InlineKeyboardButton(text="◀️ Отмена", callback_data="menu_reports")])

    return InlineKeyboardMarkup(inline_keyboard=rows)

# ── Запуск выбора периода ─────────────────────────────────────
@router.callback_query(F.data == "rep_period")
async def rep_period(call: CallbackQuery, state: FSMContext):
    now = datetime.now()
    await state.set_state(ReportStates.picking_start)
    await state.update_data(report_type="period")
    await call.message.edit_text(
        "📊 <b>Отчёт за период</b>\n\nВыберите <b>начальную</b> дату:",
        parse_mode="HTML",
        reply_markup=make_calendar(now.year, now.month, "start")
    )

@router.callback_query(F.data == "rep_growth")
async def rep_growth(call: CallbackQuery, state: FSMContext):
    now = datetime.now()
    await state.set_state(ReportStates.picking_start)
    await state.update_data(report_type="growth")
    await call.message.edit_text(
        "📈 <b>Прирост просмотров</b>\n\nВыберите <b>начальную</b> дату:",
        parse_mode="HTML",
        reply_markup=make_calendar(now.year, now.month, "start")
    )

# ── Навигация по месяцам ──────────────────────────────────────
@router.callback_query(F.data.startswith("start_nav_"))
async def start_nav(call: CallbackQuery, state: FSMContext):
    _, _, _, year, month = call.data.split("_")
    await call.message.edit_reply_markup(
        reply_markup=make_calendar(int(year), int(month), "start")
    )

@router.callback_query(F.data.startswith("end_nav_"))
async def end_nav(call: CallbackQuery, state: FSMContext):
    _, _, _, year, month = call.data.split("_")
    data = await state.get_data()
    await call.message.edit_reply_markup(
        reply_markup=make_calendar(int(year), int(month), "end", data.get("start_date"))
    )

# ── Быстрый выбор периода ─────────────────────────────────────
@router.callback_query(F.data.startswith("start_quick_"))
async def start_quick(call: CallbackQuery, state: FSMContext, pool):
    days = int(call.data.split("_")[-1])
    now  = datetime.now(tz=timezone.utc)
    date_from = now - timedelta(days=days)
    date_to   = now

    data = await state.get_data()
    report_type = data.get("report_type", "period")
    await state.clear()

    await call.message.edit_text(
        f"⏳ Формируем отчёт за последние {days} дней..."
    )
    await run_report(call, pool, report_type, date_from, date_to)

# ── Выбор начальной даты ──────────────────────────────────────
@router.callback_query(F.data.startswith("start_day_"))
async def start_day(call: CallbackQuery, state: FSMContext):
    start_date = call.data.replace("start_day_", "")
    await state.update_data(start_date=start_date)
    await state.set_state(ReportStates.picking_end)

    now = datetime.now()
    await call.message.edit_text(
        f"📅 Начало: <b>{start_date}</b>\n\nТеперь выберите <b>конечную</b> дату:",
        parse_mode="HTML",
        reply_markup=make_calendar(now.year, now.month, "end", start_date)
    )

# ── Выбор конечной даты ───────────────────────────────────────
@router.callback_query(F.data.startswith("end_day_"))
async def end_day(call: CallbackQuery, state: FSMContext, pool):
    end_date = call.data.replace("end_day_", "")
    data     = await state.get_data()
    start_date  = data.get("start_date")
    report_type = data.get("report_type", "period")

    if end_date < start_date:
        await call.answer("⚠️ Конечная дата раньше начальной!", show_alert=True)
        return

    await state.clear()
    date_from = datetime.fromisoformat(start_date).replace(tzinfo=timezone.utc)
    date_to   = datetime.fromisoformat(end_date).replace(hour=23, minute=59, second=59, tzinfo=timezone.utc)

    await call.message.edit_text("⏳ Формируем отчёт...")
    await run_report(call, pool, report_type, date_from, date_to)

@router.callback_query(F.data.startswith("end_quick_"))
async def end_quick(call: CallbackQuery, state: FSMContext, pool):
    days = int(call.data.split("_")[-1])
    now  = datetime.now(tz=timezone.utc)
    date_from = now - timedelta(days=days)
    date_to   = now

    data = await state.get_data()
    report_type = data.get("report_type", "period")
    await state.clear()

    await call.message.edit_text(f"⏳ Формируем отчёт за последние {days} дней...")
    await run_report(call, pool, report_type, date_from, date_to)

@router.callback_query(F.data == "ignore")
async def ignore(call: CallbackQuery):
    await call.answer()

# ── Запуск нужного отчёта ─────────────────────────────────────
async def run_report(call, pool, report_type, date_from, date_to):
    user_id  = call.from_user.id
    from_str = date_from.strftime("%d.%m.%Y")
    to_str   = date_to.strftime("%d.%m.%Y")

    if report_type == "period":
        text, excel = await build_period_report(pool, user_id, date_from, date_to)
        fname = f"report_{date_from.strftime('%Y%m%d')}_{date_to.strftime('%Y%m%d')}.xlsx"
    else:
        text, excel = await build_growth_report(pool, user_id, date_from, date_to)
        fname = f"growth_{date_from.strftime('%Y%m%d')}_{date_to.strftime('%Y%m%d')}.xlsx"

    await call.message.edit_text(
        text, parse_mode="HTML",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="◀️ К отчётам", callback_data="menu_reports")]
        ])
    )
    if excel:
        excel.name = fname
        await call.message.answer_document(
            document=excel,
            caption=f"📊 {from_str} — {to_str}"
        )

# ── Отчёт за период ───────────────────────────────────────────
async def build_period_report(pool, user_id, date_from, date_to):
    async with pool.acquire() as conn:
        rows = await conn.fetch("""
            SELECT
                a.article,
                COUNT(DISTINCT v.id)      AS video_count,
                COALESCE(SUM(v.views), 0) AS total_views,
                STRING_AGG(DISTINCT acc.username, ', ') AS accounts
            FROM articles a
            JOIN article_matches am ON am.article_id = a.id
            JOIN videos v   ON v.id = am.video_id
            JOIN accounts acc ON acc.id = v.account_id
            WHERE a.user_id=$1
              AND v.published_at >= $2
              AND v.published_at <= $3
            GROUP BY a.article
            ORDER BY total_views DESC
        """, user_id, date_from, date_to)

    from_str = date_from.strftime("%d.%m.%Y")
    to_str   = date_to.strftime("%d.%m.%Y")

    if not rows:
        return (
            f"📊 <b>{from_str} — {to_str}</b>\n\n"
            "Видео с вашими артикулами за этот период не найдено.\n\n"
            "<i>Убедитесь что аккаунты синхронизированы.</i>",
            None
        )

    total_videos = sum(r["video_count"] for r in rows)
    total_views  = sum(r["total_views"] for r in rows)

    lines = [
        f"📊 <b>{from_str} — {to_str}</b>\n",
        f"Видео: <b>{total_videos}</b> | Просмотры: <b>{total_views:,}</b>\n",
    ]
    for r in rows[:20]:
        lines.append(
            f"• <code>{r['article']}</code> — "
            f"{r['video_count']} видео, {r['total_views']:,} просмотров"
        )
    if len(rows) > 20:
        lines.append(f"\n<i>...ещё {len(rows)-20} артикулов в Excel</i>")

    return "\n".join(lines), build_excel_period(rows, date_from, date_to)

# ── Прирост просмотров ────────────────────────────────────────
async def build_growth_report(pool, user_id, date_from, date_to):
    async with pool.acquire() as conn:
        rows = await conn.fetch("""
            WITH
            snap_end AS (
                SELECT DISTINCT ON (s.video_id) s.video_id, s.views
                FROM snapshots s WHERE s.taken_at <= $3
                ORDER BY s.video_id, s.taken_at DESC
            ),
            snap_start AS (
                SELECT DISTINCT ON (s.video_id) s.video_id, s.views
                FROM snapshots s WHERE s.taken_at < $2
                ORDER BY s.video_id, s.taken_at DESC
            )
            SELECT
                a.article,
                COUNT(DISTINCT v.id) AS video_count,
                COALESCE(SUM(se.views - COALESCE(ss.views, 0)), 0) AS growth
            FROM articles a
            JOIN article_matches am ON am.article_id = a.id
            JOIN videos v ON v.id = am.video_id
            LEFT JOIN snap_end   se ON se.video_id = v.id
            LEFT JOIN snap_start ss ON ss.video_id = v.id
            WHERE a.user_id = $1
            GROUP BY a.article
            HAVING COALESCE(SUM(se.views - COALESCE(ss.views, 0)), 0) > 0
            ORDER BY growth DESC
        """, user_id, date_from, date_to)

    from_str = date_from.strftime("%d.%m.%Y")
    to_str   = date_to.strftime("%d.%m.%Y")

    if not rows:
        return (
            f"📈 <b>{from_str} — {to_str}</b>\n\n"
            "Прирост не зафиксирован.\n\n"
            "<i>Нужно хотя бы два замера просмотров в этом периоде.\n"
            "Бот делает замеры каждые 6 часов.</i>",
            None
        )

    total_growth = sum(r["growth"] for r in rows)
    lines = [
        f"📈 <b>{from_str} — {to_str}</b>\n",
        f"Суммарный прирост: <b>+{total_growth:,}</b> просмотров\n",
    ]
    for r in rows[:20]:
        lines.append(
            f"• <code>{r['article']}</code> — "
            f"+{r['growth']:,} просмотров ({r['video_count']} видео)"
        )
    if len(rows) > 20:
        lines.append(f"\n<i>...ещё {len(rows)-20} артикулов в Excel</i>")

    return "\n".join(lines), build_excel_growth(rows, date_from, date_to)

# ── Excel ─────────────────────────────────────────────────────
def make_excel_header(ws, headers, widths, color):
    hfill = PatternFill("solid", fgColor=color)
    hfont = Font(bold=True, color="FFFFFF")
    ctr   = Alignment(horizontal="center", vertical="center")
    for c, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font=hfont; cell.fill=hfill; cell.alignment=ctr
        ws.column_dimensions[cell.column_letter].width = w
    ws.row_dimensions[1].height = 22

def build_excel_period(rows, date_from, date_to) -> io.BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Отчёт"
    ws["A1"] = f"Отчёт {date_from.strftime('%d.%m.%Y')} — {date_to.strftime('%d.%m.%Y')}"
    ws.append([])
    make_excel_header(ws, ["Артикул","Видео","Просмотры","Аккаунты"], [20,10,15,30], "2D6A9F")
    # Сдвигаем данные на 3 строку
    wb2 = openpyxl.Workbook()
    ws2 = wb2.active
    ws2.title = "Отчёт"
    make_excel_header(ws2, ["Артикул","Кол-во видео","Просмотры","Аккаунты"], [22,14,15,30], "2D6A9F")
    altf = PatternFill("solid", fgColor="EBF3FB")
    for ri, r in enumerate(rows, 2):
        fill = altf if ri%2==0 else None
        for ci, v in enumerate([r["article"], r["video_count"], r["total_views"], r["accounts"]], 1):
            c = ws2.cell(row=ri, column=ci, value=v)
            if fill: c.fill = fill
    buf = io.BytesIO()
    wb2.save(buf); buf.seek(0)
    return buf

def build_excel_growth(rows, date_from, date_to) -> io.BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Прирост"
    make_excel_header(ws, ["Артикул","Кол-во видео","Прирост просмотров"], [22,14,20], "1E6E42")
    altf = PatternFill("solid", fgColor="E8F5E9")
    for ri, r in enumerate(rows, 2):
        fill = altf if ri%2==0 else None
        for ci, v in enumerate([r["article"], r["video_count"], r["growth"]], 1):
            c = ws.cell(row=ri, column=ci, value=v)
            if fill: c.fill = fill
    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return buf

# ── Для авто-отчётов ──────────────────────────────────────────
async def build_summary_text(pool, user_id: int, period: str) -> str:
    now = datetime.now(tz=timezone.utc)
    if period == "day":
        date_from = now - timedelta(days=1); label = "за последние сутки"
    elif period == "week":
        date_from = now - timedelta(weeks=1); label = "за последнюю неделю"
    else:
        date_from = now.replace(day=1, hour=0, minute=0, second=0); label = f"за {now.strftime('%B')}"

    async with pool.acquire() as conn:
        rows = await conn.fetch("""
            SELECT acc.username,
                COUNT(DISTINCT v.id) as videos,
                COALESCE(SUM(v.views), 0) as views
            FROM accounts acc
            JOIN videos v ON v.account_id = acc.id
            WHERE acc.user_id=$1 AND v.published_at >= $2
            GROUP BY acc.username ORDER BY views DESC
        """, user_id, date_from)

    if not rows:
        return f"🔔 Авто-отчёт {label}:\nВидео не найдено."

    lines = [f"🔔 <b>Авто-отчёт {label}</b>\n"]
    for r in rows:
        lines.append(f"@{r['username']}: {r['videos']} видео, {r['views']:,} просмотров")
    return "\n".join(lines)
